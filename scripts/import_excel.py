#!/usr/bin/env python3
import argparse, json
from pathlib import Path
from openpyxl import load_workbook

parser=argparse.ArgumentParser(description='Extract Perla product spreadsheet for review')
parser.add_argument('--file',required=True)
args=parser.parse_args()
wb=load_workbook(args.file,data_only=True,read_only=True)
ws=wb['المنتجات'] if 'المنتجات' in wb.sheetnames else wb[wb.sheetnames[0]]
products=[]
for row in ws.iter_rows(min_row=3,values_only=True):
    if not row or row[0] is None or row[1] is None or row[2] is None:
        continue
    products.append({
      'source_id':int(row[0]), 'product_code':str(row[1]), 'name_ar':str(row[2]).strip(),
      'price_sar':float(row[3] or 0), 'cost':float(row[4] or 0), 'status':str(row[5] or '')
    })
out=Path('data/catalog-import-raw.json')
out.write_text(json.dumps({'source':Path(args.file).name,'products':products},ensure_ascii=False,indent=2),encoding='utf-8')
print(f'extracted {len(products)} rows to {out}. Review before updating data/catalog.json.')
